#!/usr/bin/env python3
"""
upload_accounts.py — import an accounts spreadsheet (.xlsx or .csv) into
the Supabase `accounts` table.

Usage:
    python upload_accounts.py path/to/accounts.xlsx

Required environment variables:
    SUPABASE_URL              e.g. https://poswjjugynmxavjwfnjh.supabase.co
    SUPABASE_SERVICE_ROLE_KEY the SECRET service-role key (Project Settings
                              → API Keys → "service_role"). This bypasses
                              RLS. Never commit it.

Column detection is case-insensitive. Accepted names:
    email column:    email | e-mail | id | user | username
    password column: password | pwd | pass

Rows with a duplicate email are skipped (existing status/password kept).
Safe to re-run on the same file.
"""

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

EMAIL_KEYS = ("email", "e-mail", "id", "user", "username")
PASSWORD_KEYS = ("password", "pwd", "pass")
BATCH_SIZE = 500


def read_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook  # lazy import — not needed for CSV

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(iterator)]
        for row in iterator:
            yield dict(zip(header, row))
    elif suffix == ".csv":
        with path.open(newline="") as f:
            for row in csv.DictReader(f):
                yield row
    else:
        raise SystemExit(f"Unsupported file extension: {suffix}")


def pick_column(header, candidates):
    lowered = {k.strip().lower(): k for k in header if k}
    for candidate in candidates:
        if candidate in lowered:
            return lowered[candidate]
    return None


def normalize(rows):
    rows = list(rows)
    if not rows:
        return []
    header = list(rows[0].keys())
    email_col = pick_column(header, EMAIL_KEYS)
    pw_col = pick_column(header, PASSWORD_KEYS)
    if not email_col or not pw_col:
        raise SystemExit(
            f"Could not find email/password columns in header: {header}\n"
            f"Recognised email names: {EMAIL_KEYS}\n"
            f"Recognised password names: {PASSWORD_KEYS}"
        )

    seen = set()
    out = []
    skipped_blank = 0
    for row in rows:
        email = str(row.get(email_col) or "").strip()
        pw = str(row.get(pw_col) or "").strip()
        if not email or not pw:
            skipped_blank += 1
            continue
        if email in seen:
            continue
        seen.add(email)
        out.append({"email": email, "password": pw})

    if skipped_blank:
        print(f"Skipped {skipped_blank} row(s) with blank email/password.")
    return out


def upload(rows, base_url, service_key):
    endpoint = f"{base_url.rstrip('/')}/rest/v1/accounts?on_conflict=email"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        # ignore-duplicates: skip rows whose email already exists.
        # return=representation: response body contains the rows actually
        # inserted, so len(response) is our new-row count.
        "Prefer": "resolution=ignore-duplicates,return=representation",
    }

    inserted = 0
    total_batches = (len(rows) + BATCH_SIZE - 1) // BATCH_SIZE
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(batch).encode(),
            headers=headers,
            method="POST",
        )
        try:
            with urllib.request.urlopen(req) as resp:
                body = resp.read()
                new_rows = json.loads(body) if body else []
                inserted += len(new_rows)
                print(
                    f"  batch {i // BATCH_SIZE + 1}/{total_batches}: "
                    f"{len(new_rows)} new / {len(batch)} sent"
                )
        except urllib.error.HTTPError as e:
            detail = e.read().decode(errors="replace")
            print(f"HTTP {e.code} on batch {i // BATCH_SIZE + 1}: {detail}", file=sys.stderr)
            sys.exit(1)
    return inserted


def main():
    parser = argparse.ArgumentParser(description="Upload accounts from .xlsx/.csv into Supabase.")
    parser.add_argument("path", type=Path, help=".xlsx or .csv file to upload")
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"File not found: {args.path}")

    base_url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not key:
        raise SystemExit(
            "Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment variables "
            "before running. The service role key is the SECRET one from "
            "Project Settings → API Keys."
        )

    rows = normalize(read_rows(args.path))
    print(f"Loaded {len(rows)} unique account(s) from {args.path.name}.")
    if not rows:
        return

    inserted = upload(rows, base_url, key)
    skipped = len(rows) - inserted
    print(f"Done — {inserted} inserted, {skipped} already existed.")


if __name__ == "__main__":
    main()
